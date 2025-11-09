from sqlalchemy.orm import Session
from models import Project, Indicator, Financement
from typing import List
from decimal import Decimal
from datetime import datetime
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def generate_pdf_report(projects: List[Project], db: Session) -> bytes:
    """Generate PDF report of projects"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a73e8'),
        spaceAfter=30,
    )
    
    # Title
    elements.append(Paragraph("Rapport des Projets - ImpactTracker", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Projects data
    for project in projects:
        # Project header
        elements.append(Paragraph(f"<b>{project.titre}</b>", styles['Heading2']))
        elements.append(Spacer(1, 0.1*inch))
        
        # Project details
        data = [
            ['Domaine', project.domaine.value],
            ['Localisation', project.localisation],
            ['Pays', project.pays],
            ['Date de début', str(project.date_debut)],
            ['Budget', f"{project.budget:,.2f} EUR"],
            ['Statut', project.statut.value],
        ]
        
        table = Table(data, colWidths=[2*inch, 4*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def generate_excel_report(projects: List[Project], db: Session) -> bytes:
    """Generate Excel report of projects"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Projets"
    
    # Headers
    headers = ['ID', 'Titre', 'Domaine', 'Localisation', 'Pays', 'Date début', 'Budget', 'Statut']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for project in projects:
        ws.append([
            project.id,
            project.titre,
            project.domaine.value,
            project.localisation,
            project.pays,
            project.date_debut,
            float(project.budget),
            project.statut.value
        ])
    
    # Adjust column widths
    column_widths = [8, 30, 15, 20, 15, 12, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

